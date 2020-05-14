#! python3
# Takes number N and M from the command line and insert M blank rows at row N in spreadsheet.
# Note: in real life applications I suggest using the following function instead:
# openpyxl.worksheet.worksheet.Worksheet.insert_rows()

r"""How to use from the run window:
1.  Save this code as "C:\Users\user\PycharmProjects\HelloWorld\13 blankRowInserter.py"
2.  create 'zblankRowInserter.bat' in 'C:\Windows' with the following text content:
    @py.exe "C:\Users\user\PycharmProjects\HelloWorld\13 blankRowInserter.py" %*
3.  type in run window: 'zblankRowInserter.bat [N] [M] [filename]'
Note: There's an example 'myProduce.xlsx' on my GitHub to use for testing your own, as the book doesn't provide it.
Note: You will have to change the current working directory to the directory you stored the excel you want to use."""

import openpyxl
from sys import argv
from os import chdir
from pathlib import Path

chdir(Path.home()/r'Py3Projects\13-Working with Excel Spreadsheets')

N = argv[1]
N = int(N)
M = argv[2]
M = int(M)
filename = argv[3]

wb = openpyxl.load_workbook(filename)
sheet = wb.active

# Move cells beyond blank row to their new location.
for row in range(sheet.max_row + 1, N, -1):  # loop trough the rows backwards to avoid cell overlap
    for column in range(1, sheet.max_column + 1):
        old_location = sheet.cell(row=row, column=column)
        new_location = sheet.cell(row=row + M, column=column)
        new_location.value = old_location.value

# Clear old values from blank rows.
for row in range(N, N + M):
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=row, column=column).value = ''

wb.save(filename[:-5] + ' with blank rows' + filename[-5:])

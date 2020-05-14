#! python3
# Read content of excel file and turn it into text files.
# Note: I'm using some random poems I found online for this project

from os import chdir
from pathlib import Path
import openpyxl

chdir(Path.cwd()/r"Py3Projects\13-Working with Excel Spreadsheets")

# Create folder to put the text files in
Path(Path.cwd()/'Spreadsheet to Text Files').mkdir(exist_ok=True)

wb = openpyxl.load_workbook('Text Files to Spreadsheet.xlsx')
sheet = wb.active

# Read the content of the spreadsheet and add it to the text files:
for column in range(1, sheet.max_column + 1):
    textfile = open(Path.cwd()/rf'Spreadsheet to Text Files/poem{column}.txt', 'w')
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value is not None:
            textfile.write(cell_value)
    textfile.close()

#! python3
# Corrects costs in produce sales spreadsheet.

"""Your program does the following:
Loops over all the rows
If the row is for garlic, celery, or lemons, changes the price
This means your code will need to do the following:

Open the spreadsheet file.
For each row, check whether the value in column A is Celery, Garlic, or Lemon.
If it is, update the price in column B.
Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet, just in case)."""

from os import chdir
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

chdir(Path.home()/r'Py3Projects\automate_online-materials')

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The product type and their prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=1).font = Font(bold=True)
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('../13-Working with Excel Spreadsheets/updatedProduceSales.xlsx')

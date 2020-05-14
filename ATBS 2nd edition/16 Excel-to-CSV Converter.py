#! python3
# Converts Excel files to CSV.

"""Find all excel files
Loop over every sheet
Loop over all cells
Write to CSV file"""

import csv
import os
import openpyxl

os.chdir(r'C:\Users\Beheerder\Py3Projects\16 Working with CSV files and JSON data\Excel-to-CSV Converter')

for excel_file in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excel_file.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excel_file)

    # Loop through every sheet in the workbook.
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Create the CSV filename from the Excel filename and sheet title.
        CSV_filename = f"{excel_file}_{sheet_name}.csv"

        # Create the csv.writer object for this CSV file.
        csv_file = open(CSV_filename, 'w', newline='')
        csv_writer = csv.writer(csv_file)

        # Loop through every row in the sheet.
        for row in range(1, sheet.max_row + 1):
            row_data = []    # append each cell to this list
            # Loop through each cell in the row.
            for column in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                row_data.append(sheet.cell(row=row, column=column).value)

            csv_writer.writerow(row_data)
            # Write the rowData list to the CSV file.

        csv_file.close()

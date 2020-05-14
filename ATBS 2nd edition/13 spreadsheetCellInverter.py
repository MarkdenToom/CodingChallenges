#! python3
# Take a sheet and transpose the content into another excel sheet.
from typing import List, Any

import openpyxl
from os import chdir
from pathlib import Path

chdir(Path.home()/r'Py3Projects\13-Working with Excel Spreadsheets')

wb = openpyxl.load_workbook('soldTally.xlsx')
sheet = wb.active

# Create a nested list with structure 'sheetData[x][y]':
sheetData = []
for x in range(1, sheet.max_column + 1):
    column = []
    for y in range(1, sheet.max_row + 1):
        column.append(sheet.cell(row=y, column=x).value)
    sheetData.append(column)

# Add the data to a new workbook:
wb = openpyxl.Workbook()
sheet = wb.active
for row in range(1, len(sheetData) + 1):
    for column in range(1, len(sheetData[0]) + 1):
        sheet.cell(row=row, column=column).value = sheetData[row - 1][column - 1]

wb.save('soldTally - Transposed.xlsx')

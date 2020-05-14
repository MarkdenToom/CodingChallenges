#! python3
# Takes number N from the command line and creates an N*N multiplication table in excel.

r"""How to use:
1.  Save this code as "C:\Users\user\PycharmProjects\HelloWorld\13 multiplicationTable.py"
2.  create 'zmultiplicationTable.bat' in 'C:\Windows' with the following line:
    @py.exe "C:\Users\Beheerder\PycharmProjects\HelloWorld\13 multiplicationTable.py" %*
3.  type in run (Uitvoeren) window: 'zmultiplicationTable.bat [N]'"""

import openpyxl
from sys import argv
from os import chdir
from pathlib import Path

chdir(Path.home()/r'Py3Projects\13-Working with Excel Spreadsheets')

N = int(argv[1])
wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, N + 1):
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=i + 1, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=1, column=i + 1).font = openpyxl.styles.Font(bold=True)

for row in range(2, N + 2):
    for column in range(2, N + 2):
        sheet.cell(row=row, column=column).value = (row - 1) * (column - 1)

wb.save('multiplicationTable.xlsx')
